#!/usr/bin/env python3
"""
Script to create Partner Dashboard Test Cases Excel file
Extracts test cases from documentation and creates a comprehensive Excel workbook
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Define test cases data structure
test_cases = []

# ==================== DASHBOARD TEST CASES ====================
# UI Test Cases
test_cases.extend([
    {
        "Test Case ID": "UI-DASH-001",
        "Module": "Dashboard",
        "Test Case Name": "Header Title Display",
        "Test Type": "UI",
        "Priority": "Low",
        "Preconditions": "User logged in, Dashboard page loaded",
        "Test Steps": "1. Navigate to Dashboard\n2. Verify header title displays 'ダッシュボード'",
        "Expected Results": "Header title is visible and displays correct Japanese text",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "UI-DASH-002",
        "Module": "Dashboard",
        "Test Case Name": "Unprocessed Orders Metric Display",
        "Test Type": "UI",
        "Priority": "High",
        "Preconditions": "User logged in, Dashboard page loaded",
        "Test Steps": "1. Navigate to Dashboard\n2. Locate Unprocessed Orders metric card\n3. Verify number is displayed\n4. Verify card is clickable",
        "Expected Results": "Metric card shows unprocessed orders count, card is clickable",
        "Status": "Pass",
        "Notes": "Currently showing 6 unprocessed orders"
    },
    {
        "Test Case ID": "UI-DASH-003",
        "Module": "Dashboard",
        "Test Case Name": "Sales Today Metric Display",
        "Test Type": "UI",
        "Priority": "High",
        "Preconditions": "User logged in, Dashboard page loaded",
        "Test Steps": "1. Navigate to Dashboard\n2. Locate Sales Today metric card\n3. Verify currency format (¥) is displayed",
        "Expected Results": "Metric card shows sales amount in Japanese yen format",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "UI-DASH-004",
        "Module": "Dashboard",
        "Test Case Name": "Order Count Metric Display",
        "Test Type": "UI",
        "Priority": "Medium",
        "Preconditions": "User logged in, Dashboard page loaded",
        "Test Steps": "1. Navigate to Dashboard\n2. Locate Order Count metric card\n3. Verify number is displayed",
        "Expected Results": "Metric card shows order count as a number",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "UI-DASH-005",
        "Module": "Dashboard",
        "Test Case Name": "Quick Action: Create Order Button",
        "Test Type": "UI",
        "Priority": "High",
        "Preconditions": "User logged in, Dashboard page loaded",
        "Test Steps": "1. Navigate to Dashboard\n2. Locate 'Create Order' quick action button\n3. Verify button is visible and clickable",
        "Expected Results": "Button is visible and clickable",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "UI-DASH-006",
        "Module": "Dashboard",
        "Test Case Name": "Quick Action: Add Customer Button",
        "Test Type": "UI",
        "Priority": "Medium",
        "Preconditions": "User logged in, Dashboard page loaded",
        "Test Steps": "1. Navigate to Dashboard\n2. Locate 'Add Customer' quick action button\n3. Verify button is visible and clickable",
        "Expected Results": "Button is visible and clickable",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "UI-DASH-007",
        "Module": "Dashboard",
        "Test Case Name": "Sidebar Navigation Highlight",
        "Test Type": "UI",
        "Priority": "Low",
        "Preconditions": "User logged in, Dashboard page loaded",
        "Test Steps": "1. Navigate to Dashboard\n2. Verify sidebar navigation highlights 'Dashboard' menu item",
        "Expected Results": "Dashboard menu item is highlighted in sidebar",
        "Status": "Pass",
        "Notes": ""
    }
])

# API Test Cases
test_cases.extend([
    {
        "Test Case ID": "API-DASH-001",
        "Module": "Dashboard",
        "Test Case Name": "Get Dashboard Metrics",
        "Test Type": "API",
        "Priority": "Critical",
        "Preconditions": "User authenticated, valid session token",
        "Test Steps": "1. Send GET request to /api/dashboard/metrics\n2. Verify response status code\n3. Verify response contains sales, orders, unprocessed counts",
        "Expected Results": "Returns 200 OK with metrics data (sales, orders, unprocessed counts)",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "API-DASH-002",
        "Module": "Dashboard",
        "Test Case Name": "Get Dashboard Chart Data",
        "Test Type": "API",
        "Priority": "Medium",
        "Preconditions": "User authenticated, valid session token",
        "Test Steps": "1. Send GET request to /api/dashboard/chart\n2. Verify response status code\n3. Verify response contains chart data",
        "Expected Results": "Returns 200 OK with data for sales chart",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "API-DASH-003",
        "Module": "Dashboard",
        "Test Case Name": "Get System Notifications",
        "Test Type": "API",
        "Priority": "Medium",
        "Preconditions": "User authenticated, valid session token",
        "Test Steps": "1. Send GET request to /api/notifications\n2. Verify response status code\n3. Verify response contains recent notifications",
        "Expected Results": "Returns 200 OK with recent system notifications",
        "Status": "Pending",
        "Notes": ""
    }
])

# Functional Test Cases
test_cases.extend([
    {
        "Test Case ID": "TC-DASH-001",
        "Module": "Dashboard",
        "Test Case Name": "Dashboard Page Load",
        "Test Type": "Functional",
        "Priority": "Critical",
        "Preconditions": "User is logged in, Valid session token",
        "Test Steps": "1. Navigate to /\n2. Wait for page to load",
        "Expected Results": "Dashboard loads within 3 seconds, All 7 metric cards displayed, User information correct, Latest activity shows recent orders, Quick actions visible, No console errors",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-DASH-002",
        "Module": "Dashboard",
        "Test Case Name": "Metric Values Display",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "Dashboard is loaded, Metrics data available",
        "Test Steps": "1. Verify Today's Orders value\n2. Verify Today's Order Amount value\n3. Verify Today's Sales value\n4. Verify Monthly Order Amount value\n5. Verify Monthly Sales value\n6. Verify New Customers value\n7. Verify Unprocessed Orders value",
        "Expected Results": "All metrics display correct values, Currency formatted as ¥X,XXX,XXX, Numbers formatted with suffix (件, 人), Warning badge shows on Unprocessed Orders if > 0",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-DASH-003",
        "Module": "Dashboard",
        "Test Case Name": "Latest Activity Display",
        "Test Type": "Functional",
        "Priority": "Medium",
        "Preconditions": "Dashboard is loaded, Recent orders exist",
        "Test Steps": "1. Scroll to Latest Activity section\n2. Verify activity items displayed\n3. Check order numbers correct\n4. Verify dates are recent\n5. Click on activity item",
        "Expected Results": "At least 3-5 recent activities shown, Order numbers in format #XXXX, Dates in YYYY-MM-DD format, Clicking navigates to order detail page",
        "Status": "Pending",
        "Notes": "Requires clicking activity item"
    },
    {
        "Test Case ID": "TC-DASH-004",
        "Module": "Dashboard",
        "Test Case Name": "Quick Actions Functionality",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "Dashboard is loaded, User has appropriate permissions",
        "Test Steps": "1. Click 'Create New Customer' button\n2. Verify navigation to customer creation page\n3. Go back to dashboard\n4. Click 'Create New Order' button\n5. Verify navigation to order creation page\n6. Repeat for all 5 quick actions",
        "Expected Results": "Each button navigates to correct page, No errors during navigation, User can return to dashboard",
        "Status": "Pending",
        "Notes": "Requires navigation testing"
    },
    {
        "Test Case ID": "TC-DASH-005",
        "Module": "Dashboard",
        "Test Case Name": "Responsive Design",
        "Test Type": "Functional",
        "Priority": "Medium",
        "Preconditions": "Dashboard is loaded",
        "Test Steps": "1. Resize browser to desktop width (1920px)\n2. Verify all 7 metrics in one row\n3. Resize to tablet width (768px)\n4. Verify metrics wrap to 3-4 columns\n5. Resize to mobile width (375px)\n6. Verify metrics stack vertically",
        "Expected Results": "Layout adapts to screen size, No horizontal scrolling, All content remains accessible, Sidebar collapses on mobile",
        "Status": "Pending",
        "Notes": "Requires responsive testing"
    },
    {
        "Test Case ID": "TC-DASH-006",
        "Module": "Dashboard",
        "Test Case Name": "Auto-Refresh Functionality",
        "Test Type": "Functional",
        "Priority": "Low",
        "Preconditions": "Dashboard is loaded, Auto-refresh enabled",
        "Test Steps": "1. Note current metric values\n2. Wait for auto-refresh interval (30-60 seconds)\n3. Verify metrics update automatically\n4. Check for API calls in network tab",
        "Expected Results": "Metrics refresh automatically, No page reload required, API calls made at regular intervals, UI updates smoothly without flicker",
        "Status": "Pending",
        "Notes": "Requires time-based testing"
    },
    {
        "Test Case ID": "TC-DASH-007",
        "Module": "Dashboard",
        "Test Case Name": "Error Handling - API Failure",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "Dashboard is loading",
        "Test Steps": "1. Simulate API failure (disconnect network)\n2. Attempt to load dashboard\n3. Verify error message displayed\n4. Reconnect network\n5. Click retry button",
        "Expected Results": "Error message displayed to user, Retry button available, Dashboard loads successfully after retry, No data loss or corruption",
        "Status": "Pending",
        "Notes": "Requires error simulation"
    },
    {
        "Test Case ID": "TC-DASH-008",
        "Module": "Dashboard",
        "Test Case Name": "Permission-Based Display",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "User logged in with specific role",
        "Test Steps": "1. Login as Store Owner\n2. Verify all metrics visible\n3. Verify all quick actions visible\n4. Login as different role (if available)\n5. Verify appropriate restrictions",
        "Expected Results": "Store Owner sees all metrics, Mall Administrator sees all features, Other roles see limited features (if applicable), No unauthorized access to data",
        "Status": "Pending",
        "Notes": "Requires multi-role testing"
    },
    {
        "Test Case ID": "TC-DASH-009",
        "Module": "Dashboard",
        "Test Case Name": "404 API Errors",
        "Test Type": "Functional",
        "Priority": "Medium",
        "Preconditions": "Dashboard is loading",
        "Test Steps": "1. Open browser console\n2. Load dashboard\n3. Check for 404 errors\n4. Verify errors for: /system/stores, /system/yakatas, /system/locations",
        "Expected Results": "ISSUE FOUND: 404 errors present, Dashboard still functions despite errors, No impact on user experience, Errors should be fixed or calls removed",
        "Status": "Fail",
        "Notes": "Issue Identified - 404 errors for system endpoints"
    },
    {
        "Test Case ID": "TC-DASH-010",
        "Module": "Dashboard",
        "Test Case Name": "Data Accuracy",
        "Test Type": "Functional",
        "Priority": "Critical",
        "Preconditions": "Known order data in system",
        "Test Steps": "1. Manually count orders in Shopify for today\n2. Compare with Dashboard 'Today's Orders'\n3. Calculate total order amount for month\n4. Compare with Dashboard 'Monthly Order Amount'\n5. Verify unprocessed orders count",
        "Expected Results": "Dashboard metrics match Shopify data, Calculations are accurate, No data discrepancies, Real-time sync working",
        "Status": "Pending",
        "Notes": "Requires Shopify access for verification"
    }
])

# ==================== ORDER MANAGEMENT TEST CASES ====================
# UI Test Cases
test_cases.extend([
    {
        "Test Case ID": "UI-ORD-001",
        "Module": "Order Management",
        "Test Case Name": "Order Table Display",
        "Test Type": "UI",
        "Priority": "Critical",
        "Preconditions": "User logged in, Navigate to /orders",
        "Test Steps": "1. Navigate to Order Management\n2. Verify order table is visible\n3. Verify table has 9 columns",
        "Expected Results": "Order table is visible with 9 columns (Checkbox, Status, Date, Customer, Order Number, Amount, Shipments, Elapsed Days, Actions)",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "UI-ORD-002",
        "Module": "Order Management",
        "Test Case Name": "Search Bar Display",
        "Test Type": "UI",
        "Priority": "High",
        "Preconditions": "User logged in, Navigate to /orders",
        "Test Steps": "1. Navigate to Order Management\n2. Locate search bar\n3. Verify placeholder text is correct",
        "Expected Results": "Search bar is visible with placeholder: '注文番号、顧客名、メールアドレス、電話番号、配送先住所で検索'",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "UI-ORD-003",
        "Module": "Order Management",
        "Test Case Name": "Status Filter Dropdown",
        "Test Type": "UI",
        "Priority": "High",
        "Preconditions": "User logged in, Navigate to /orders",
        "Test Steps": "1. Navigate to Order Management\n2. Locate status filter dropdown\n3. Verify dropdown contains all statuses",
        "Expected Results": "Status filter dropdown is visible and contains all order statuses (Draft, Processing, Completed, Pending, etc.)",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "UI-ORD-004",
        "Module": "Order Management",
        "Test Case Name": "Pagination Controls",
        "Test Type": "UI",
        "Priority": "Medium",
        "Preconditions": "User logged in, Navigate to /orders",
        "Test Steps": "1. Navigate to Order Management\n2. Locate pagination controls\n3. Verify 'Next/Prev' buttons are working",
        "Expected Results": "Pagination controls are visible with working Next/Previous buttons",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "UI-ORD-005",
        "Module": "Order Management",
        "Test Case Name": "Create Order Button",
        "Test Type": "UI",
        "Priority": "High",
        "Preconditions": "User logged in, Navigate to /orders",
        "Test Steps": "1. Navigate to Order Management\n2. Locate 'Create Order' button\n3. Verify button is visible and clickable",
        "Expected Results": "'Create Order' button is visible and clickable",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "UI-ORD-006",
        "Module": "Order Management",
        "Test Case Name": "Bulk Action Menu",
        "Test Type": "UI",
        "Priority": "Medium",
        "Preconditions": "User logged in, Navigate to /orders",
        "Test Steps": "1. Navigate to Order Management\n2. Verify bulk action menu is hidden\n3. Select multiple rows\n4. Verify bulk action menu appears",
        "Expected Results": "Bulk action menu is hidden until rows are selected, Menu appears when rows are selected",
        "Status": "Pass",
        "Notes": ""
    }
])

# API Test Cases
test_cases.extend([
    {
        "Test Case ID": "API-ORD-001",
        "Module": "Order Management",
        "Test Case Name": "Get Orders List",
        "Test Type": "API",
        "Priority": "Critical",
        "Preconditions": "User authenticated, Valid session token",
        "Test Steps": "1. Send GET request to /api/orders\n2. Verify response status code\n3. Verify response contains paginated order list",
        "Expected Results": "Returns 200 OK with paginated order list",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "API-ORD-002",
        "Module": "Order Management",
        "Test Case Name": "Get Order Details",
        "Test Type": "API",
        "Priority": "Critical",
        "Preconditions": "User authenticated, Valid order ID",
        "Test Steps": "1. Send GET request to /api/orders/{id}\n2. Verify response status code\n3. Verify response contains order details",
        "Expected Results": "Returns 200 OK with single order details",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "API-ORD-003",
        "Module": "Order Management",
        "Test Case Name": "Update Order Status",
        "Test Type": "API",
        "Priority": "High",
        "Preconditions": "User authenticated, Valid order ID",
        "Test Steps": "1. Send PATCH request to /api/orders/{id}/status\n2. Verify response status code\n3. Verify order status is updated",
        "Expected Results": "Returns 200 OK, Order status updated successfully",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "API-ORD-004",
        "Module": "Order Management",
        "Test Case Name": "Search Orders",
        "Test Type": "API",
        "Priority": "High",
        "Preconditions": "User authenticated",
        "Test Steps": "1. Send GET request to /api/orders/search?q=query\n2. Verify response status code\n3. Verify response contains search results",
        "Expected Results": "Returns 200 OK with search results matching query",
        "Status": "Pass",
        "Notes": "Partial search may fail"
    }
])

# Functional Test Cases
test_cases.extend([
    {
        "Test Case ID": "TC-ORD-001",
        "Module": "Order Management",
        "Test Case Name": "Order List Page Load",
        "Test Type": "Functional",
        "Priority": "Critical",
        "Preconditions": "User is logged in, Orders exist in system",
        "Test Steps": "1. Navigate to /orders\n2. Wait for page to load",
        "Expected Results": "Page loads within 3 seconds, Table displays 10 orders (first page), Pagination shows '1–10 of 177', All columns visible, Search bar present, Action buttons visible",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-ORD-002",
        "Module": "Order Management",
        "Test Case Name": "Search Functionality",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "Order list is loaded",
        "Test Steps": "1. Click on search field\n2. Type 'customer@example.com'\n3. Wait for search results",
        "Expected Results": "Search executes after debounce (300ms), Results filter to matching orders, Pagination updates to reflect filtered results, Clear search button (X) appears",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-ORD-003",
        "Module": "Order Management",
        "Test Case Name": "Pagination Navigation",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "Order list loaded with 177 orders",
        "Test Steps": "1. Verify current page shows '1–10 of 177'\n2. Click 'Next' button (>)\n3. Verify page shows '11–20 of 177'\n4. Click 'Previous' button (<)\n5. Verify back to '1–10 of 177'",
        "Expected Results": "Next button navigates to page 2, Previous button navigates to page 1, URL updates with ?page parameter, Table data updates, Previous disabled on page 1, Next disabled on last page",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-ORD-004",
        "Module": "Order Management",
        "Test Case Name": "Rows Per Page Change",
        "Test Type": "Functional",
        "Priority": "Medium",
        "Preconditions": "Order list loaded",
        "Test Steps": "1. Click rows per page dropdown\n2. Select '25'\n3. Verify table updates",
        "Expected Results": "Table shows 25 rows, Pagination updates to '1–25 of 177', Total pages recalculated, Preference saved (optional)",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-ORD-005",
        "Module": "Order Management",
        "Test Case Name": "Order Row Click Navigation",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "Order list loaded",
        "Test Steps": "1. Click on first order row\n2. Verify navigation to order detail",
        "Expected Results": "Navigates to /orders/{orderId}, Order detail page loads, Breadcrumb updates, Order information displayed correctly",
        "Status": "Pass",
        "Notes": "Verified via screenshot"
    },
    {
        "Test Case ID": "TC-ORD-006",
        "Module": "Order Management",
        "Test Case Name": "Bulk Selection",
        "Test Type": "Functional",
        "Priority": "Medium",
        "Preconditions": "Order list loaded",
        "Test Steps": "1. Click checkbox on first order\n2. Verify selection count shows '1 selected'\n3. Click checkbox on second order\n4. Verify count shows '2 selected'\n5. Click 'Select All' checkbox\n6. Verify all 10 visible orders selected",
        "Expected Results": "Individual selection works, Selection count updates, Select all selects all visible rows, Bulk action bar appears when items selected",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-ORD-007",
        "Module": "Order Management",
        "Test Case Name": "Column Sorting",
        "Test Type": "Functional",
        "Priority": "Medium",
        "Preconditions": "Order list loaded",
        "Test Steps": "1. Click on 'Date' column header\n2. Verify orders sort by date ascending\n3. Click 'Date' again\n4. Verify orders sort by date descending\n5. Repeat for 'Amount' column",
        "Expected Results": "Clicking column header toggles sort, Sort icon appears in header, Data reorders correctly, URL updates with sort parameters",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-ORD-008",
        "Module": "Order Management",
        "Test Case Name": "Order Status Display",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "Orders with different statuses exist",
        "Test Steps": "1. Verify status badges are color-coded\n2. Check each status type: Draft (Gray), Pending (Orange), Processing (Blue), Completed (Green), Refunded (Red)",
        "Expected Results": "Each status has distinct color, Badge is readable, Color matches design system, Status text is in Japanese",
        "Status": "Pass",
        "Notes": "Verified via screenshot"
    },
    {
        "Test Case ID": "TC-ORD-009",
        "Module": "Order Management",
        "Test Case Name": "Action Menu Functionality",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "Order list loaded",
        "Test Steps": "1. Click actions menu (⋮) on first order\n2. Verify menu options appear\n3. Click 'View Details'\n4. Verify navigation to detail page",
        "Expected Results": "Menu opens on click, All options visible (View Details, Edit Order, Process Refund, Cancel Order, Print Invoice, Send Email), Clicking option performs action",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-ORD-010",
        "Module": "Order Management",
        "Test Case Name": "Responsive Design",
        "Test Type": "Functional",
        "Priority": "Medium",
        "Preconditions": "Order list loaded",
        "Test Steps": "1. Resize browser to tablet width (768px)\n2. Verify table adapts\n3. Resize to mobile width (375px)\n4. Verify mobile layout",
        "Expected Results": "Table scrolls horizontally on small screens, Important columns remain visible, Action buttons stack vertically on mobile, Search bar remains full width",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-ORD-011",
        "Module": "Order Management",
        "Test Case Name": "Order Detail Page Load",
        "Test Type": "Functional",
        "Priority": "Critical",
        "Preconditions": "User navigated to order detail",
        "Test Steps": "1. Verify order header displays\n2. Check customer information card\n3. Verify order items table\n4. Check shipping information\n5. Verify payment information\n6. Check order timeline",
        "Expected Results": "All sections load correctly, Data matches order, No missing information, Action buttons available",
        "Status": "Pass",
        "Notes": "Verified via screenshot"
    },
    {
        "Test Case ID": "TC-ORD-012",
        "Module": "Order Management",
        "Test Case Name": "Shopify Data Sync",
        "Test Type": "Functional",
        "Priority": "Critical",
        "Preconditions": "Shopify webhook configured",
        "Test Steps": "1. Create new order in Shopify\n2. Wait for webhook to trigger\n3. Refresh order list in Partner Dashboard\n4. Verify new order appears",
        "Expected Results": "Order syncs within 5 seconds, All data correctly mapped, Status matches Shopify, Customer info accurate",
        "Status": "Pending",
        "Notes": "Requires Shopify access"
    }
])

# Continue with remaining modules... (Customer, Account, System, Login, E2E, Integration)
# Due to length, I'll add the most critical ones and create the script structure

def create_excel_file():
    """Create Excel workbook with test cases"""
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create main test cases sheet
    ws = wb.create_sheet("Test Cases", 0)
    
    # Define headers
    headers = [
        "Test Case ID", "Module", "Test Case Name", "Test Type", 
        "Priority", "Preconditions", "Test Steps", "Expected Results", 
        "Status", "Notes"
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Write test cases
    for row_num, test_case in enumerate(test_cases, 2):
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            value = test_case.get(header, "")
            cell.value = value
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Color code by priority
            if header == "Priority":
                if value == "Critical":
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                elif value == "High":
                    cell.fill = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                elif value == "Medium":
                    cell.fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
                elif value == "Low":
                    cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            
            # Color code by status
            if header == "Status":
                if value == "Pass":
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                elif value == "Fail":
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                elif value == "Pending":
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                elif value == "Blocked":
                    cell.fill = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")
    
    # Set column widths
    column_widths = {
        "A": 15,  # Test Case ID
        "B": 20,  # Module
        "C": 40,  # Test Case Name
        "D": 12,  # Test Type
        "E": 12,  # Priority
        "F": 30,  # Preconditions
        "G": 50,  # Test Steps
        "H": 50,  # Expected Results
        "I": 12,  # Status
        "J": 40   # Notes
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Create summary sheet
    create_summary_sheet(wb)
    
    # Save workbook
    filename = "Partner Dashboard Test Cases.xlsx"
    wb.save(filename)
    print(f"Excel file created: {filename}")
    print(f"Total test cases: {len(test_cases)}")
    
    return filename

def create_summary_sheet(wb):
    """Create summary sheet with statistics"""
    ws = wb.create_sheet("Summary", 1)
    
    # Title
    ws['A1'] = "Partner Dashboard Test Cases - Summary"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')
    
    # Statistics by Module
    ws['A3'] = "Test Cases by Module"
    ws['A3'].font = Font(bold=True, size=12)
    
    modules = {}
    for tc in test_cases:
        module = tc.get("Module", "Unknown")
        modules[module] = modules.get(module, 0) + 1
    
    row = 4
    ws.cell(row=row, column=1).value = "Module"
    ws.cell(row=row, column=2).value = "Count"
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    row = 5
    for module, count in sorted(modules.items()):
        ws.cell(row=row, column=1).value = module
        ws.cell(row=row, column=2).value = count
        row += 1
    
    # Statistics by Priority
    row += 2
    ws.cell(row=row, column=1).value = "Test Cases by Priority"
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    
    priorities = {}
    for tc in test_cases:
        priority = tc.get("Priority", "Unknown")
        priorities[priority] = priorities.get(priority, 0) + 1
    
    row += 1
    ws.cell(row=row, column=1).value = "Priority"
    ws.cell(row=row, column=2).value = "Count"
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    row += 1
    for priority in ["Critical", "High", "Medium", "Low"]:
        if priority in priorities:
            ws.cell(row=row, column=1).value = priority
            ws.cell(row=row, column=2).value = priorities[priority]
            row += 1
    
    # Statistics by Status
    row += 2
    ws.cell(row=row, column=1).value = "Test Cases by Status"
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    
    statuses = {}
    for tc in test_cases:
        status = tc.get("Status", "Unknown")
        statuses[status] = statuses.get(status, 0) + 1
    
    row += 1
    ws.cell(row=row, column=1).value = "Status"
    ws.cell(row=row, column=2).value = "Count"
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    row += 1
    for status in ["Pass", "Fail", "Pending", "Blocked"]:
        if status in statuses:
            ws.cell(row=row, column=1).value = status
            ws.cell(row=row, column=2).value = statuses[status]
            row += 1
    
    # Statistics by Test Type
    row += 2
    ws.cell(row=row, column=1).value = "Test Cases by Type"
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    
    types = {}
    for tc in test_cases:
        test_type = tc.get("Test Type", "Unknown")
        types[test_type] = types.get(test_type, 0) + 1
    
    row += 1
    ws.cell(row=row, column=1).value = "Test Type"
    ws.cell(row=row, column=2).value = "Count"
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    row += 1
    for test_type, count in sorted(types.items()):
        ws.cell(row=row, column=1).value = test_type
        ws.cell(row=row, column=2).value = count
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15

# ==================== CUSTOMER MANAGEMENT TEST CASES ====================
# UI Test Cases
test_cases.extend([
    {
        "Test Case ID": "UI-CUST-001",
        "Module": "Customer Management",
        "Test Case Name": "Customer Table Display",
        "Test Type": "UI",
        "Priority": "Critical",
        "Preconditions": "User logged in, Navigate to /customers",
        "Test Steps": "1. Navigate to Customer Management\n2. Verify customer table is visible\n3. Verify table has 8 columns",
        "Expected Results": "Customer table is visible with 8 columns (Customer ID, Name, Email, Phone, Orders, Total Spent, Last Order, Status, Actions)",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "UI-CUST-002",
        "Module": "Customer Management",
        "Test Case Name": "Search Bar Display",
        "Test Type": "UI",
        "Priority": "High",
        "Preconditions": "User logged in, Navigate to /customers",
        "Test Steps": "1. Navigate to Customer Management\n2. Locate search bar\n3. Verify placeholder text",
        "Expected Results": "Search bar is visible with placeholder: 'Name, Email, Phone'",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "UI-CUST-003",
        "Module": "Customer Management",
        "Test Case Name": "Import CSV Button",
        "Test Type": "UI",
        "Priority": "Medium",
        "Preconditions": "User logged in, Navigate to /customers",
        "Test Steps": "1. Navigate to Customer Management\n2. Locate 'Import CSV' button\n3. Verify button is visible and clickable",
        "Expected Results": "'Import CSV' button is visible and clickable",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "UI-CUST-004",
        "Module": "Customer Management",
        "Test Case Name": "Customer Detail Header",
        "Test Type": "UI",
        "Priority": "High",
        "Preconditions": "User logged in, Customer detail page open",
        "Test Steps": "1. Open customer detail page\n2. Verify header displays Name, Email, ID",
        "Expected Results": "Customer detail header shows Name, Email, and Customer ID",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "UI-CUST-005",
        "Module": "Customer Management",
        "Test Case Name": "Order History Section",
        "Test Type": "UI",
        "Priority": "High",
        "Preconditions": "User logged in, Customer detail page open",
        "Test Steps": "1. Open customer detail page\n2. Verify Order History section is visible",
        "Expected Results": "Order History section is visible in Detail view",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "UI-CUST-006",
        "Module": "Customer Management",
        "Test Case Name": "Notes Section",
        "Test Type": "UI",
        "Priority": "Medium",
        "Preconditions": "User logged in, Customer detail page open",
        "Test Steps": "1. Open customer detail page\n2. Verify Notes section is visible and editable",
        "Expected Results": "Notes section is visible and editable",
        "Status": "Pass",
        "Notes": ""
    }
])

# API Test Cases
test_cases.extend([
    {
        "Test Case ID": "API-CUST-001",
        "Module": "Customer Management",
        "Test Case Name": "Get Customers List",
        "Test Type": "API",
        "Priority": "Critical",
        "Preconditions": "User authenticated, Valid session token",
        "Test Steps": "1. Send GET request to /api/customers\n2. Verify response status code\n3. Verify response contains paginated customer list",
        "Expected Results": "Returns 200 OK with paginated customer list",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "API-CUST-002",
        "Module": "Customer Management",
        "Test Case Name": "Get Customer Details",
        "Test Type": "API",
        "Priority": "Critical",
        "Preconditions": "User authenticated, Valid customer ID",
        "Test Steps": "1. Send GET request to /api/customers/{id}\n2. Verify response status code\n3. Verify response contains customer profile & history",
        "Expected Results": "Returns 200 OK with customer profile and order history",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "API-CUST-003",
        "Module": "Customer Management",
        "Test Case Name": "CSV Import",
        "Test Type": "API",
        "Priority": "Medium",
        "Preconditions": "User authenticated, Valid CSV file",
        "Test Steps": "1. Send POST request to /api/customers/import with CSV file\n2. Verify response status code\n3. Verify customers are imported",
        "Expected Results": "Returns 200 OK, Handles CSV upload successfully",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "API-CUST-004",
        "Module": "Customer Management",
        "Test Case Name": "Add Customer Note",
        "Test Type": "API",
        "Priority": "Medium",
        "Preconditions": "User authenticated, Valid customer ID",
        "Test Steps": "1. Send POST request to /api/customers/{id}/notes\n2. Verify response status code\n3. Verify note is added",
        "Expected Results": "Returns 201 Created, Note added to customer",
        "Status": "Pending",
        "Notes": ""
    }
])

# Functional Test Cases
test_cases.extend([
    {
        "Test Case ID": "TC-CUST-001",
        "Module": "Customer Management",
        "Test Case Name": "Customer List Page Load",
        "Test Type": "Functional",
        "Priority": "Critical",
        "Preconditions": "User is logged in, Customers exist in system",
        "Test Steps": "1. Navigate to /customers\n2. Wait for page to load",
        "Expected Results": "Page loads with customer table, search bar, and pagination",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-CUST-002",
        "Module": "Customer Management",
        "Test Case Name": "Search Functionality",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "Customer list is loaded",
        "Test Steps": "1. Type customer name in search\n2. Verify results filter",
        "Expected Results": "Results filter to matching customers",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-CUST-003",
        "Module": "Customer Management",
        "Test Case Name": "Customer Detail Navigation",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "Customer list is loaded",
        "Test Steps": "1. Click on customer name\n2. Verify navigation to customer detail page",
        "Expected Results": "Navigates to customer detail page",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-CUST-004",
        "Module": "Customer Management",
        "Test Case Name": "CSV Import",
        "Test Type": "Functional",
        "Priority": "Medium",
        "Preconditions": "User logged in, Valid CSV file available",
        "Test Steps": "1. Navigate to CSV Import page\n2. Upload valid CSV file\n3. Verify import process",
        "Expected Results": "Customers imported successfully",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-CUST-005",
        "Module": "Customer Management",
        "Test Case Name": "Shopify Data Sync",
        "Test Type": "Functional",
        "Priority": "Critical",
        "Preconditions": "Shopify webhook configured",
        "Test Steps": "1. Create customer in Shopify\n2. Wait for sync\n3. Verify customer appears in Partner Dashboard",
        "Expected Results": "Customer appears in Partner Dashboard",
        "Status": "Pending",
        "Notes": "Requires Shopify access"
    },
    {
        "Test Case ID": "TC-CUST-006",
        "Module": "Customer Management",
        "Test Case Name": "Customer Status Display",
        "Test Type": "Functional",
        "Priority": "Medium",
        "Preconditions": "Customer list loaded",
        "Test Steps": "1. Verify status badges are color-coded correctly",
        "Expected Results": "Status badges color-coded correctly (Active=Green, Inactive=Gray, VIP=Gold)",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-CUST-007",
        "Module": "Customer Management",
        "Test Case Name": "Order History Display",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "Customer detail page open",
        "Test Steps": "1. View customer detail page\n2. Verify order history table shows customer's orders",
        "Expected Results": "Order history table shows customer's orders",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-CUST-008",
        "Module": "Customer Management",
        "Test Case Name": "Customer Edit",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "Customer detail page open",
        "Test Steps": "1. Click edit\n2. Modify customer info\n3. Save",
        "Expected Results": "Customer updated successfully",
        "Status": "Pending",
        "Notes": ""
    }
])

# ==================== ACCOUNT MANAGEMENT TEST CASES ====================
# UI Test Cases
test_cases.extend([
    {
        "Test Case ID": "UI-ACC-001",
        "Module": "Account Management",
        "Test Case Name": "Account Table Display",
        "Test Type": "UI",
        "Priority": "Critical",
        "Preconditions": "User logged in as Mall Admin, Navigate to /accounts",
        "Test Steps": "1. Navigate to Account Management\n2. Verify account table is visible\n3. Verify table has 7 columns",
        "Expected Results": "Account table is visible with 7 columns (Employee Code, Account Name, Status, Email, Permissions, Creation Date, Actions)",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "UI-ACC-002",
        "Module": "Account Management",
        "Test Case Name": "Create Account Button",
        "Test Type": "UI",
        "Priority": "Critical",
        "Preconditions": "User logged in as Mall Admin, Navigate to /accounts",
        "Test Steps": "1. Navigate to Account Management\n2. Verify 'Create Account' button is visible (Admin only)",
        "Expected Results": "'Create Account' button is visible (Admin only)",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "UI-ACC-003",
        "Module": "Account Management",
        "Test Case Name": "Role Selector",
        "Test Type": "UI",
        "Priority": "High",
        "Preconditions": "User logged in as Mall Admin, Account creation form open",
        "Test Steps": "1. Open account creation form\n2. Locate role selector dropdown\n3. Verify options (Admin, Owner, Manager, Staff)",
        "Expected Results": "Role selector dropdown contains options: Admin, Owner, Manager, Staff",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "UI-ACC-004",
        "Module": "Account Management",
        "Test Case Name": "Permission Matrix",
        "Test Type": "UI",
        "Priority": "High",
        "Preconditions": "User logged in as Mall Admin, Account detail/edit view open",
        "Test Steps": "1. Open account detail/edit view\n2. Verify permission matrix is visible",
        "Expected Results": "Permission matrix is visible in Detail/Edit view",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "UI-ACC-005",
        "Module": "Account Management",
        "Test Case Name": "Status Toggle",
        "Test Type": "UI",
        "Priority": "High",
        "Preconditions": "User logged in as Mall Admin, Account detail page open",
        "Test Steps": "1. Open account detail page\n2. Locate status toggle (Active/Inactive switch)\n3. Verify toggle is functional",
        "Expected Results": "Status toggle (Active/Inactive switch) is visible and functional",
        "Status": "Pass",
        "Notes": ""
    }
])

# API Test Cases
test_cases.extend([
    {
        "Test Case ID": "API-ACC-001",
        "Module": "Account Management",
        "Test Case Name": "Get Accounts List",
        "Test Type": "API",
        "Priority": "Critical",
        "Preconditions": "User authenticated as Mall Admin, Valid session token",
        "Test Steps": "1. Send GET request to /api/accounts\n2. Verify response status code\n3. Verify response contains account list",
        "Expected Results": "Returns 200 OK with account list",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "API-ACC-002",
        "Module": "Account Management",
        "Test Case Name": "Create Account",
        "Test Type": "API",
        "Priority": "Critical",
        "Preconditions": "User authenticated as Mall Admin, Valid account data",
        "Test Steps": "1. Send POST request to /api/accounts with account data\n2. Verify response status code\n3. Verify account is created",
        "Expected Results": "Returns 201 Created, Account created successfully",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "API-ACC-003",
        "Module": "Account Management",
        "Test Case Name": "Update Permissions",
        "Test Type": "API",
        "Priority": "High",
        "Preconditions": "User authenticated as Mall Admin, Valid account ID",
        "Test Steps": "1. Send PATCH request to /api/accounts/{id}/permissions\n2. Verify response status code\n3. Verify permissions are updated",
        "Expected Results": "Returns 200 OK, Permissions updated successfully",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "API-ACC-004",
        "Module": "Account Management",
        "Test Case Name": "Update Account Status",
        "Test Type": "API",
        "Priority": "High",
        "Preconditions": "User authenticated as Mall Admin, Valid account ID",
        "Test Steps": "1. Send PATCH request to /api/accounts/{id}/status\n2. Verify response status code\n3. Verify account status is updated",
        "Expected Results": "Returns 200 OK, Account status updated (Deactivates/Activates account)",
        "Status": "Pending",
        "Notes": ""
    }
])

# Functional Test Cases
test_cases.extend([
    {
        "Test Case ID": "TC-ACC-001",
        "Module": "Account Management",
        "Test Case Name": "Account List Page Load",
        "Test Type": "Functional",
        "Priority": "Critical",
        "Preconditions": "User logged in as Mall Admin",
        "Test Steps": "1. Navigate to /accounts\n2. Wait for page to load",
        "Expected Results": "Page loads with account table (Mall Admin only)",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-ACC-002",
        "Module": "Account Management",
        "Test Case Name": "Access Control Verification",
        "Test Type": "Functional",
        "Priority": "Critical",
        "Preconditions": "User logged in as Store Owner (Non-Admin)",
        "Test Steps": "1. Login as non-admin user\n2. Attempt to access /accounts",
        "Expected Results": "Access Denied / Redirect to Dashboard (403 Forbidden or redirect)",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-ACC-003",
        "Module": "Account Management",
        "Test Case Name": "Create Account",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "User logged in as Mall Admin, Navigate to /accounts",
        "Test Steps": "1. Click Create Account\n2. Fill form with valid details\n3. Submit",
        "Expected Results": "Success message displayed, New account appears in list, Temporary password generated",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-ACC-004",
        "Module": "Account Management",
        "Test Case Name": "Update Permissions",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "User logged in as Mall Admin, Account detail page open",
        "Test Steps": "1. Open Account Detail\n2. Change permissions (e.g., uncheck 'Order Management')\n3. Save",
        "Expected Results": "Permissions updated, User session invalidated, User can no longer access Orders module (verify with login)",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-ACC-005",
        "Module": "Account Management",
        "Test Case Name": "Deactivate Account",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "User logged in as Mall Admin, Account detail page open",
        "Test Steps": "1. Click deactivate on active account\n2. Toggle status to 'Inactive'",
        "Expected Results": "Account status changes to inactive, User cannot login",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-ACC-006",
        "Module": "Account Management",
        "Test Case Name": "Access History Display",
        "Test Type": "Functional",
        "Priority": "Medium",
        "Preconditions": "User logged in as Mall Admin, Account detail page open",
        "Test Steps": "1. View account detail page\n2. Navigate to Access History section",
        "Expected Results": "Access history table shows login attempts with IP addresses and devices",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-ACC-007",
        "Module": "Account Management",
        "Test Case Name": "Role-Based Permissions",
        "Test Type": "Functional",
        "Priority": "Critical",
        "Preconditions": "User logged in as Mall Admin",
        "Test Steps": "1. Create accounts with different roles\n2. Verify each role has appropriate default permissions",
        "Expected Results": "Each role has appropriate default permissions",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-ACC-008",
        "Module": "Account Management",
        "Test Case Name": "Location-Based Access",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "User logged in as Mall Admin",
        "Test Steps": "1. Assign specific store/building/location to account\n2. Login as that account\n3. Verify access restrictions",
        "Expected Results": "User can only access assigned locations",
        "Status": "Pending",
        "Notes": ""
    }
])

# ==================== SYSTEM MANAGEMENT TEST CASES ====================
# UI Test Cases
test_cases.extend([
    {
        "Test Case ID": "UI-SYS-001",
        "Module": "System Management",
        "Test Case Name": "Log Table Display",
        "Test Type": "UI",
        "Priority": "High",
        "Preconditions": "User logged in as Mall Admin, Navigate to /system/logs",
        "Test Steps": "1. Navigate to System Logs\n2. Verify log table is visible\n3. Verify table shows Timestamp, User, Action columns",
        "Expected Results": "Log table is visible showing Timestamp, User, Action",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "UI-SYS-002",
        "Module": "System Management",
        "Test Case Name": "Date Range Filter",
        "Test Type": "UI",
        "Priority": "High",
        "Preconditions": "User logged in as Mall Admin, Navigate to /system/logs",
        "Test Steps": "1. Navigate to System Logs\n2. Locate date range filter\n3. Verify date pickers are working",
        "Expected Results": "Date range filter is visible with working date pickers",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "UI-SYS-003",
        "Module": "System Management",
        "Test Case Name": "Module Filter",
        "Test Type": "UI",
        "Priority": "Medium",
        "Preconditions": "User logged in as Mall Admin, Navigate to /system/logs",
        "Test Steps": "1. Navigate to System Logs\n2. Locate module filter dropdown\n3. Verify dropdown is working",
        "Expected Results": "Module filter dropdown is visible and working",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "UI-SYS-004",
        "Module": "System Management",
        "Test Case Name": "Integration Settings Display",
        "Test Type": "UI",
        "Priority": "High",
        "Preconditions": "User logged in as Mall Admin, Navigate to /system/integrations",
        "Test Steps": "1. Navigate to Integration Settings\n2. Verify settings are visible (RITS only currently)",
        "Expected Results": "Integration settings visible (RITS only currently, Shopify settings NOT displayed - Known Issue)",
        "Status": "Fail",
        "Notes": "Known Issue: Shopify settings not displayed"
    },
    {
        "Test Case ID": "UI-SYS-005",
        "Module": "System Management",
        "Test Case Name": "Sidebar Sub-menus",
        "Test Type": "UI",
        "Priority": "Medium",
        "Preconditions": "User logged in as Mall Admin, Navigate to /system",
        "Test Steps": "1. Navigate to System Management\n2. Verify sidebar sub-menus are expandable (Logs, Stores, Locations)",
        "Expected Results": "Sidebar sub-menus are expandable (Logs, Stores, Locations)",
        "Status": "Pass",
        "Notes": ""
    }
])

# API Test Cases
test_cases.extend([
    {
        "Test Case ID": "API-SYS-001",
        "Module": "System Management",
        "Test Case Name": "Get System Logs",
        "Test Type": "API",
        "Priority": "High",
        "Preconditions": "User authenticated as Mall Admin, Valid session token",
        "Test Steps": "1. Send GET request to /api/system/logs\n2. Verify response status code\n3. Verify response contains system logs",
        "Expected Results": "Returns 200 OK with system logs",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "API-SYS-002",
        "Module": "System Management",
        "Test Case Name": "Get Integration Settings",
        "Test Type": "API",
        "Priority": "High",
        "Preconditions": "User authenticated as Mall Admin, Valid session token",
        "Test Steps": "1. Send GET request to /api/system/integrations\n2. Verify response status code\n3. Verify response contains integration config",
        "Expected Results": "Returns 200 OK with integration config",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "API-SYS-003",
        "Module": "System Management",
        "Test Case Name": "Get Stores (404 Error)",
        "Test Type": "API",
        "Priority": "High",
        "Preconditions": "User authenticated, Valid session token",
        "Test Steps": "1. Send GET request to /api/system/stores\n2. Verify response status code",
        "Expected Results": "Returns 404 Not Found (Known Bug - Should return stores)",
        "Status": "Fail",
        "Notes": "Known Bug: 404 error for /system/stores endpoint"
    }
])

# Functional Test Cases
test_cases.extend([
    {
        "Test Case ID": "TC-SYS-001",
        "Module": "System Management",
        "Test Case Name": "Log Management Page Load",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "User logged in as Mall Admin",
        "Test Steps": "1. Navigate to /system/logs\n2. Wait for page to load",
        "Expected Results": "Page loads with log table and filters (access restricted to Mall Admin)",
        "Status": "Pending",
        "Notes": "Access restricted to Mall Admin"
    },
    {
        "Test Case ID": "TC-SYS-002",
        "Module": "System Management",
        "Test Case Name": "Log Filtering by Date",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "User logged in as Mall Admin, Log management page loaded",
        "Test Steps": "1. Select date range\n2. Click Apply",
        "Expected Results": "Logs filtered to date range",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-SYS-003",
        "Module": "System Management",
        "Test Case Name": "Log Filtering by Module",
        "Test Type": "Functional",
        "Priority": "Medium",
        "Preconditions": "User logged in as Mall Admin, Log management page loaded",
        "Test Steps": "1. Select module from dropdown\n2. Apply filter",
        "Expected Results": "Only logs from selected module shown",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-SYS-004",
        "Module": "System Management",
        "Test Case Name": "Log Export",
        "Test Type": "Functional",
        "Priority": "Medium",
        "Preconditions": "User logged in as Mall Admin, Log management page loaded",
        "Test Steps": "1. Apply filters\n2. Click Export\n3. Select CSV format",
        "Expected Results": "CSV file downloads with filtered logs",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-SYS-005",
        "Module": "System Management",
        "Test Case Name": "Store Creation",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "User logged in as Mall Admin, Navigate to Store Management",
        "Test Steps": "1. Navigate to Store Management\n2. Create new store\n3. Fill store details\n4. Submit",
        "Expected Results": "Store created successfully",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-SYS-006",
        "Module": "System Management",
        "Test Case Name": "Location Hierarchy",
        "Test Type": "Functional",
        "Priority": "Critical",
        "Preconditions": "User logged in as Mall Admin",
        "Test Steps": "1. Create Store\n2. Create Building under Store\n3. Create Location under Building",
        "Expected Results": "Hierarchical structure maintained (Store > Building > Location)",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-SYS-007",
        "Module": "System Management",
        "Test Case Name": "Shopify Integration Test",
        "Test Type": "Functional",
        "Priority": "Critical",
        "Preconditions": "User logged in as Mall Admin, Navigate to Integration Settings",
        "Test Steps": "1. Enter Shopify credentials\n2. Test connection\n3. Verify webhooks configured",
        "Expected Results": "Connection successful, Webhooks configured",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-SYS-008",
        "Module": "System Management",
        "Test Case Name": "Access Control Verification",
        "Test Type": "Functional",
        "Priority": "Critical",
        "Preconditions": "User logged in as non-admin",
        "Test Steps": "1. Login as non-admin\n2. Attempt to access /system",
        "Expected Results": "403 Forbidden or redirect",
        "Status": "Pending",
        "Notes": ""
    }
])

# ==================== LOGIN/AUTHENTICATION TEST CASES ====================
test_cases.extend([
    {
        "Test Case ID": "TC-LOGIN-001",
        "Module": "Login/Authentication",
        "Test Case Name": "Verify Store Dropdown Functionality",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "Navigate to login page, Page fully loaded",
        "Test Steps": "1. Locate the Store dropdown (storeCode)\n2. Click on the Store dropdown\n3. Observe the available options\n4. Select 'Tokyo Main Store (TKY001)'\n5. Verify selection is displayed",
        "Expected Results": "Dropdown displays '店舗を選択してください' initially, Clicking opens dropdown with available store options, Options include 'Tokyo Main Store (TKY001)', Selected value is displayed",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-LOGIN-002",
        "Module": "Login/Authentication",
        "Test Case Name": "Verify Building Dropdown Cascading Logic",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "Navigate to login page, Store dropdown not yet selected",
        "Test Steps": "1. Observe initial state of Building dropdown\n2. Verify Building dropdown is disabled\n3. Select a store from Store dropdown\n4. Observe Building dropdown state change\n5. Click Building dropdown\n6. Verify options are loaded based on selected store\n7. Select 'Building A (TKY001-A)'",
        "Expected Results": "Building dropdown is disabled initially, Building dropdown becomes enabled after Store selection, Building options are filtered based on selected Store, Options include 'Building A (TKY001-A)', Selected value is displayed",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-LOGIN-003",
        "Module": "Login/Authentication",
        "Test Case Name": "Verify Location Dropdown Cascading Logic",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "Navigate to login page, Store and Building not yet selected",
        "Test Steps": "1. Observe initial state of Location dropdown\n2. Verify Location dropdown is disabled\n3. Select a store from Store dropdown\n4. Select a building from Building dropdown\n5. Observe Location dropdown state change\n6. Click Location dropdown\n7. Verify options are loaded based on selected building",
        "Expected Results": "Location dropdown is disabled initially, Location dropdown remains disabled after only Store selection, Location dropdown becomes enabled after Building selection, Location options are filtered based on selected Building",
        "Status": "Pass",
        "Notes": "Location dropdown opens and displays options correctly, but automated selection has UI interaction challenges"
    },
    {
        "Test Case ID": "TC-LOGIN-004",
        "Module": "Login/Authentication",
        "Test Case Name": "Verify Email/Employee Code Input Field",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "Navigate to login page",
        "Test Steps": "1. Locate the email/employee code input field\n2. Click on the field\n3. Enter email: 'admin@itfor-wl.myshopify.com'\n4. Verify input is displayed",
        "Expected Results": "Field accepts text input, Email is displayed in the field, No validation errors appear for valid email format",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-LOGIN-005",
        "Module": "Login/Authentication",
        "Test Case Name": "Verify Password Input Field",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "Navigate to login page",
        "Test Steps": "1. Locate the password input field\n2. Click on the field\n3. Enter password: 'Weblife_123'\n4. Verify password is masked (shows dots/asterisks)\n5. Click password visibility toggle icon\n6. Verify password becomes visible",
        "Expected Results": "Field accepts text input, Password characters are masked by default, Visibility toggle icon is present, Clicking toggle reveals/hides password",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-LOGIN-006",
        "Module": "Login/Authentication",
        "Test Case Name": "Verify Remember Me Checkbox",
        "Test Type": "Functional",
        "Priority": "Medium",
        "Preconditions": "Navigate to login page",
        "Test Steps": "1. Locate the 'ログイン情報を保存する' (Save login information) checkbox\n2. Verify checkbox is unchecked by default\n3. Click the checkbox\n4. Verify checkbox becomes checked\n5. Click again\n6. Verify checkbox becomes unchecked",
        "Expected Results": "Checkbox is unchecked by default, Clicking toggles the checked state, Visual indication of checked/unchecked state is clear",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "TC-LOGIN-007",
        "Module": "Login/Authentication",
        "Test Case Name": "Verify Login with All Fields Filled",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "Navigate to login page, Valid credentials available",
        "Test Steps": "1. Enter email: 'admin@itfor-wl.myshopify.com'\n2. Enter password: 'Weblife_123'\n3. Select Store: 'Tokyo Main Store (TKY001)'\n4. Select Building: 'Building A (TKY001-A)'\n5. Select Location: 'Floor 1 (TKY001-A-F1)'\n6. Check 'Remember Me' checkbox\n7. Click 'ログイン' (Login) button\n8. Observe result",
        "Expected Results": "All fields accept input, Login button is clickable, Form submits successfully, User is redirected to dashboard/home page, No error messages appear",
        "Status": "Blocked",
        "Notes": "Unable to complete automated test due to Location dropdown selection issue. Manual testing required."
    },
    {
        "Test Case ID": "TC-LOGIN-008",
        "Module": "Login/Authentication",
        "Test Case Name": "Verify Required Field Validation",
        "Test Type": "Functional",
        "Priority": "High",
        "Preconditions": "Navigate to login page",
        "Test Steps": "1. Observe email/employee code field label\n2. Observe password field label\n3. Verify asterisk (*) indicators are present\n4. Attempt to submit form with empty fields\n5. Observe validation messages",
        "Expected Results": "Required fields show asterisk (*) indicator, Submitting empty form shows validation errors, Error messages are clear and helpful",
        "Status": "Pending",
        "Notes": "Requires manual testing to verify validation behavior"
    },
    {
        "Test Case ID": "TC-LOGIN-009",
        "Module": "Login/Authentication",
        "Test Case Name": "Verify Dropdown Dependency Reset",
        "Test Type": "Functional",
        "Priority": "Medium",
        "Preconditions": "Navigate to login page, All dropdowns have selections",
        "Test Steps": "1. Select Store, Building, and Location\n2. Change Store selection to a different store\n3. Observe Building dropdown\n4. Observe Location dropdown\n5. Verify both are reset",
        "Expected Results": "Changing Store resets Building selection, Changing Store resets Location selection, Building dropdown shows placeholder text, Location dropdown becomes disabled",
        "Status": "Pending",
        "Notes": "Requires manual testing to verify reset behavior"
    },
    {
        "Test Case ID": "TC-LOGIN-010",
        "Module": "Login/Authentication",
        "Test Case Name": "Verify Login Button State",
        "Test Type": "Functional",
        "Priority": "Medium",
        "Preconditions": "Navigate to login page",
        "Test Steps": "1. Observe Login button with empty form\n2. Verify button is enabled\n3. Fill some fields\n4. Verify button remains enabled\n5. Fill all fields\n6. Verify button remains enabled",
        "Expected Results": "Login button is always enabled, Validation occurs on submit, not on button state",
        "Status": "Pass",
        "Notes": ""
    }
])

# ==================== E2E TEST CASES ====================
test_cases.extend([
    {
        "Test Case ID": "AUTH-001",
        "Module": "E2E - Authentication",
        "Test Case Name": "Login with Valid Credentials",
        "Test Type": "E2E",
        "Priority": "Critical",
        "Preconditions": "Valid credentials available",
        "Test Steps": "1. Enter valid email/password\n2. Select Store/Building/Location\n3. Click Login",
        "Expected Results": "Dashboard loads successfully",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "AUTH-002",
        "Module": "E2E - Authentication",
        "Test Case Name": "Login with Invalid Credentials",
        "Test Type": "E2E",
        "Priority": "High",
        "Preconditions": "Invalid password available",
        "Test Steps": "1. Enter invalid password\n2. Click Login",
        "Expected Results": "Error message displayed",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "AUTH-003",
        "Module": "E2E - Authentication",
        "Test Case Name": "RBAC Verification (Store Owner)",
        "Test Type": "E2E",
        "Priority": "High",
        "Preconditions": "Store Owner credentials available",
        "Test Steps": "1. Login as Store Owner\n2. Attempt to access Account Management",
        "Expected Results": "Access denied or restricted view",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "AUTH-004",
        "Module": "E2E - Authentication",
        "Test Case Name": "RBAC Verification (Mall Admin)",
        "Test Type": "E2E",
        "Priority": "Critical",
        "Preconditions": "Mall Admin credentials available",
        "Test Steps": "1. Login as Mall Admin\n2. Access Account Management",
        "Expected Results": "Full access granted",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "ORD-001",
        "Module": "E2E - Order Management",
        "Test Case Name": "View Order List",
        "Test Type": "E2E",
        "Priority": "High",
        "Preconditions": "User logged in",
        "Test Steps": "1. Navigate to Order Management",
        "Expected Results": "List loads with correct columns",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "ORD-002",
        "Module": "E2E - Order Management",
        "Test Case Name": "Filter Orders by Status",
        "Test Type": "E2E",
        "Priority": "Medium",
        "Preconditions": "User logged in, Order list loaded",
        "Test Steps": "1. Select 'Unprocessed' status",
        "Expected Results": "Only unprocessed orders shown",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "ORD-003",
        "Module": "E2E - Order Management",
        "Test Case Name": "Search Order by ID",
        "Test Type": "E2E",
        "Priority": "Critical",
        "Preconditions": "User logged in, Order list loaded",
        "Test Steps": "1. Enter Order ID in search",
        "Expected Results": "Specific order displayed",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "ORD-004",
        "Module": "E2E - Order Management",
        "Test Case Name": "View Order Details",
        "Test Type": "E2E",
        "Priority": "High",
        "Preconditions": "User logged in, Order list loaded",
        "Test Steps": "1. Click on order row",
        "Expected Results": "Detail page shows items, customer, shipping",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "ORD-005",
        "Module": "E2E - Order Management",
        "Test Case Name": "Update Order Status",
        "Test Type": "E2E",
        "Priority": "Critical",
        "Preconditions": "User logged in, Order detail page open",
        "Test Steps": "1. Open order\n2. Change status to 'Shipped'",
        "Expected Results": "Status updates, history logged",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "ORD-006",
        "Module": "E2E - Order Management",
        "Test Case Name": "Bulk Order Print",
        "Test Type": "E2E",
        "Priority": "Medium",
        "Preconditions": "User logged in, Order list loaded",
        "Test Steps": "1. Select multiple orders\n2. Click 'Print Form'",
        "Expected Results": "PDF generated for selected orders",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "CUST-001",
        "Module": "E2E - Customer Management",
        "Test Case Name": "Search Customer",
        "Test Type": "E2E",
        "Priority": "High",
        "Preconditions": "User logged in, Customer list loaded",
        "Test Steps": "1. Enter customer name/email",
        "Expected Results": "Correct customer listed",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "CUST-002",
        "Module": "E2E - Customer Management",
        "Test Case Name": "View Customer Profile",
        "Test Type": "E2E",
        "Priority": "High",
        "Preconditions": "User logged in, Customer list loaded",
        "Test Steps": "1. Click customer name",
        "Expected Results": "Profile shows info and order history",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "CUST-003",
        "Module": "E2E - Customer Management",
        "Test Case Name": "CSV Import",
        "Test Type": "E2E",
        "Priority": "Medium",
        "Preconditions": "User logged in, Valid CSV file available",
        "Test Steps": "1. Upload customer CSV",
        "Expected Results": "Customers created/updated in bulk",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "CUST-004",
        "Module": "E2E - Customer Management",
        "Test Case Name": "Edit Customer Notes",
        "Test Type": "E2E",
        "Priority": "Low",
        "Preconditions": "User logged in, Customer profile open",
        "Test Steps": "1. Open profile\n2. Add note",
        "Expected Results": "Note saved and visible",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "ACC-001",
        "Module": "E2E - Account Management",
        "Test Case Name": "Create New Account",
        "Test Type": "E2E",
        "Priority": "High",
        "Preconditions": "User logged in as Mall Admin",
        "Test Steps": "1. Click Create\n2. Fill details & permissions",
        "Expected Results": "Account created, email sent",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "ACC-002",
        "Module": "E2E - Account Management",
        "Test Case Name": "Edit Permissions",
        "Test Type": "E2E",
        "Priority": "Critical",
        "Preconditions": "User logged in as Mall Admin, Account detail page open",
        "Test Steps": "1. Select account\n2. Change module access",
        "Expected Results": "User access updates immediately",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "ACC-003",
        "Module": "E2E - Account Management",
        "Test Case Name": "Deactivate Account",
        "Test Type": "E2E",
        "Priority": "High",
        "Preconditions": "User logged in as Mall Admin, Account detail page open",
        "Test Steps": "1. Set status to Inactive",
        "Expected Results": "User cannot login",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "SYS-001",
        "Module": "E2E - System Management",
        "Test Case Name": "View System Logs",
        "Test Type": "E2E",
        "Priority": "Medium",
        "Preconditions": "User logged in as Mall Admin",
        "Test Steps": "1. Navigate to Logs\n2. Filter by date",
        "Expected Results": "Logs displayed correctly",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "SYS-002",
        "Module": "E2E - System Management",
        "Test Case Name": "RITS Integration Check",
        "Test Type": "E2E",
        "Priority": "High",
        "Preconditions": "User logged in as Mall Admin",
        "Test Steps": "1. Navigate to Integration\n2. Verify settings",
        "Expected Results": "Settings visible (read-only/edit)",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "SYS-003",
        "Module": "E2E - System Management",
        "Test Case Name": "Shopify Sync (Backend)",
        "Test Type": "E2E",
        "Priority": "Critical",
        "Preconditions": "Shopify webhook configured",
        "Test Steps": "1. Create order in Shopify\n2. Check Dashboard",
        "Expected Results": "Order appears in Dashboard (latency < 5m)",
        "Status": "Pending",
        "Notes": "Requires Shopify access"
    },
    {
        "Test Case ID": "E2E-001",
        "Module": "E2E - Order Fulfillment",
        "Test Case Name": "Order Fulfillment Journey",
        "Test Type": "E2E",
        "Priority": "Critical",
        "Preconditions": "User logged in as Store Owner",
        "Test Steps": "1. Login as Store Owner\n2. Dashboard -> Check 'Unprocessed Orders'\n3. Open Unprocessed Order\n4. Verify Items & Payment\n5. Print Packing Slip\n6. Update Status -> 'Processing'\n7. Update Status -> 'Shipped' (Add Tracking #)\n8. Verify Status in Order List",
        "Expected Results": "Order status updates correctly at each step",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "E2E-002",
        "Module": "E2E - Customer Support",
        "Test Case Name": "Customer Support Journey",
        "Test Type": "E2E",
        "Priority": "High",
        "Preconditions": "User logged in as Store Manager",
        "Test Steps": "1. Login as Store Manager\n2. Navigate to Customers\n3. Search for 'Tanaka'\n4. Open Profile\n5. Check Order History\n6. Add Note: 'Customer called regarding shipping delay'\n7. Save Note",
        "Expected Results": "Note is saved and visible in timeline",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "E2E-003",
        "Module": "E2E - User Onboarding",
        "Test Case Name": "User Onboarding Journey",
        "Test Type": "E2E",
        "Priority": "High",
        "Preconditions": "User logged in as Mall Admin",
        "Test Steps": "1. Login as Mall Admin\n2. Navigate to Accounts\n3. Create Account (Role: Staff, Store: Tokyo)\n4. Logout\n5. Login as New Staff\n6. Verify access (Can view Orders, Cannot view Accounts)",
        "Expected Results": "New user created with correct permissions",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "E2E-NEG-001",
        "Module": "E2E - Negative Testing",
        "Test Case Name": "Unauthorized Access",
        "Test Type": "E2E",
        "Priority": "High",
        "Preconditions": "User logged in as Staff",
        "Test Steps": "1. Login as Staff\n2. Attempt to access /accounts URL directly",
        "Expected Results": "403 Forbidden or Redirect to Dashboard",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "E2E-NEG-002",
        "Module": "E2E - Negative Testing",
        "Test Case Name": "Invalid Login",
        "Test Type": "E2E",
        "Priority": "High",
        "Preconditions": "Invalid password available",
        "Test Steps": "1. Enter valid email\n2. Enter wrong password",
        "Expected Results": "Error message 'Invalid credentials'",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "E2E-NEG-003",
        "Module": "E2E - Negative Testing",
        "Test Case Name": "404 Page Handling",
        "Test Type": "E2E",
        "Priority": "Medium",
        "Preconditions": "User logged in",
        "Test Steps": "1. Navigate to /non-existent-page",
        "Expected Results": "Custom 404 Page displayed (not generic browser error)",
        "Status": "Pending",
        "Notes": ""
    }
])

# ==================== INTEGRATION TEST CASES ====================
test_cases.extend([
    {
        "Test Case ID": "INT-SHO-001",
        "Module": "Integration - Shopify",
        "Test Case Name": "Order Sync (Shopify -> Dashboard)",
        "Test Type": "Integration",
        "Priority": "Critical",
        "Preconditions": "Shopify webhook configured, Valid Shopify store",
        "Test Steps": "1. Create order in Shopify\n2. Wait for sync interval (or trigger webhook)\n3. Check Dashboard Order List",
        "Expected Results": "Order appears with correct details (Items, Customer, Address)",
        "Status": "Pending",
        "Notes": "Requires Shopify access"
    },
    {
        "Test Case ID": "INT-SHO-002",
        "Module": "Integration - Shopify",
        "Test Case Name": "Customer Sync (Shopify -> Dashboard)",
        "Test Type": "Integration",
        "Priority": "High",
        "Preconditions": "Shopify webhook configured, Valid Shopify store",
        "Test Steps": "1. Create customer in Shopify\n2. Check Dashboard Customer List",
        "Expected Results": "Customer appears with correct contact info",
        "Status": "Pending",
        "Notes": "Requires Shopify access"
    },
    {
        "Test Case ID": "INT-SHO-003",
        "Module": "Integration - Shopify",
        "Test Case Name": "Inventory Sync (Dashboard -> Shopify)",
        "Test Type": "Integration",
        "Priority": "High",
        "Preconditions": "Shopify integration configured, Inventory management enabled",
        "Test Steps": "1. Update inventory in Dashboard (if supported)\n2. Check Shopify Product",
        "Expected Results": "Inventory count updates in Shopify",
        "Status": "Pending",
        "Notes": "Requires Shopify access"
    },
    {
        "Test Case ID": "INT-SHO-004",
        "Module": "Integration - Shopify",
        "Test Case Name": "Fulfillment Sync (Dashboard -> Shopify)",
        "Test Type": "Integration",
        "Priority": "Critical",
        "Preconditions": "Shopify integration configured, Order exists",
        "Test Steps": "1. Mark order as 'Shipped' in Dashboard\n2. Add tracking number\n3. Check Shopify Order",
        "Expected Results": "Order status becomes 'Fulfilled', tracking info added",
        "Status": "Pending",
        "Notes": "Requires Shopify access"
    },
    {
        "Test Case ID": "INT-SHO-005",
        "Module": "Integration - Shopify",
        "Test Case Name": "Order Creation Webhook",
        "Test Type": "Integration",
        "Priority": "Critical",
        "Preconditions": "Webhook endpoint configured",
        "Test Steps": "1. Simulate 'orders/create' webhook payload\n2. Send to endpoint",
        "Expected Results": "System accepts payload (200 OK), creates order",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "INT-SHO-006",
        "Module": "Integration - Shopify",
        "Test Case Name": "Order Update Webhook",
        "Test Type": "Integration",
        "Priority": "High",
        "Preconditions": "Webhook endpoint configured",
        "Test Steps": "1. Simulate 'orders/updated' webhook\n2. Send to endpoint",
        "Expected Results": "System updates existing order status/details",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "INT-SHO-007",
        "Module": "Integration - Shopify",
        "Test Case Name": "Webhook Retry Logic",
        "Test Type": "Integration",
        "Priority": "Medium",
        "Preconditions": "Webhook endpoint configured",
        "Test Steps": "1. Simulate 500 error on webhook receiver\n2. Retry payload",
        "Expected Results": "System eventually processes payload after recovery",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "INT-SHO-008",
        "Module": "Integration - Shopify",
        "Test Case Name": "Invalid API Credentials",
        "Test Type": "Integration",
        "Priority": "High",
        "Preconditions": "Integration settings accessible",
        "Test Steps": "1. Invalidate API Key in settings (Backend)\n2. Trigger sync",
        "Expected Results": "System logs authentication error, alerts admin",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "INT-SHO-009",
        "Module": "Integration - Shopify",
        "Test Case Name": "Rate Limiting",
        "Test Type": "Integration",
        "Priority": "Medium",
        "Preconditions": "Shopify API access",
        "Test Steps": "1. Trigger rapid-fire sync requests",
        "Expected Results": "System handles rate limits gracefully (429 handling)",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "INT-INT-001",
        "Module": "Integration - Internal",
        "Test Case Name": "Order-Customer Linkage",
        "Test Type": "Integration",
        "Priority": "Critical",
        "Preconditions": "User logged in, Order exists",
        "Test Steps": "1. View Order Details\n2. Click Customer Name link",
        "Expected Results": "Redirects to correct Customer Profile",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "INT-INT-002",
        "Module": "Integration - Internal",
        "Test Case Name": "Customer-Order History",
        "Test Type": "Integration",
        "Priority": "Critical",
        "Preconditions": "User logged in, Customer exists with orders",
        "Test Steps": "1. View Customer Profile\n2. Check Order History",
        "Expected Results": "Lists all orders associated with that customer",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "INT-INT-003",
        "Module": "Integration - Internal",
        "Test Case Name": "Account-Store Access",
        "Test Type": "Integration",
        "Priority": "High",
        "Preconditions": "User logged in as Mall Admin",
        "Test Steps": "1. Create Account with access to Store A only\n2. Login as Account\n3. Try to view Orders from Store B",
        "Expected Results": "Access Denied or Data Filtered (Store B orders hidden)",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "INT-RITS-001",
        "Module": "Integration - RITS",
        "Test Case Name": "Configuration Load",
        "Test Type": "Integration",
        "Priority": "High",
        "Preconditions": "User logged in as Mall Admin",
        "Test Steps": "1. Navigate to /system/integrations\n2. Check RITS settings",
        "Expected Results": "Settings load from backend",
        "Status": "Pass",
        "Notes": ""
    },
    {
        "Test Case ID": "INT-RITS-002",
        "Module": "Integration - RITS",
        "Test Case Name": "Connection Test",
        "Test Type": "Integration",
        "Priority": "High",
        "Preconditions": "User logged in as Mall Admin, RITS settings configured",
        "Test Steps": "1. Click 'Test Connection' (if available)\n2. Or trigger RITS sync",
        "Expected Results": "Successful connection log entry",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "INT-AUTH-001",
        "Module": "Integration - Authentication",
        "Test Case Name": "Session Timeout",
        "Test Type": "Integration",
        "Priority": "Medium",
        "Preconditions": "User logged in",
        "Test Steps": "1. Login\n2. Wait for timeout duration\n3. Refresh page",
        "Expected Results": "Redirected to Login page",
        "Status": "Pending",
        "Notes": ""
    },
    {
        "Test Case ID": "INT-AUTH-002",
        "Module": "Integration - Authentication",
        "Test Case Name": "Concurrent Sessions",
        "Test Type": "Integration",
        "Priority": "Low",
        "Preconditions": "User credentials available",
        "Test Steps": "1. Login on Device A\n2. Login on Device B (same user)",
        "Expected Results": "Both active OR Device A logged out (depending on policy)",
        "Status": "Pending",
        "Notes": ""
    }
])

if __name__ == "__main__":
    create_excel_file()

